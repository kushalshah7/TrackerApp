from __future__ import annotations
import os, re, shutil, tempfile
from copy import copy
from datetime import date, datetime
from pathlib import Path
from threading import RLock
from typing import Any
import openpyxl
import portalocker
from openpyxl.formula.translate import Translator
from openpyxl.styles import PatternFill
from .config import BACKUP_DIR, MODULES, STATUS_FIELDS, TEMPLATE_PATH, WORKBOOK_PATH

class WorkbookUnavailable(RuntimeError): pass

MISSING_VALUE_FILL = PatternFill(fill_type="solid", fgColor="FFFF0000")
MISSING_VALUE_COLOR = "FFFF0000"

class ExcelService:
    def __init__(self, path: Path = WORKBOOK_PATH):
        self.path = Path(path); self.lock_path = self.path.with_suffix(".lock"); self._thread_lock = RLock()

    def _ensure(self):
        if not self.path.exists():
            raise WorkbookUnavailable(f"Workbook not found. Place it at {self.path}")

    def _load(self):
        self._ensure(); return openpyxl.load_workbook(self.path)

    @staticmethod
    def _copy_style(ws, source: int, target: int, start_col: int, count: int):
        if source <= 0: return
        ws.row_dimensions[target].height = ws.row_dimensions[source].height
        for col in range(start_col, start_col + count):
            a, b = ws.cell(source, col), ws.cell(target, col)
            if a.has_style:
                b.font, b.fill, b.border = copy(a.font), copy(a.fill), copy(a.border)
                b.alignment, b.protection, b.number_format = copy(a.alignment), copy(a.protection), a.number_format

    @staticmethod
    def _headers(ws, cfg):
        return [ws.cell(cfg["header_row"], cfg["start_col"] + i).value for i in range(len(cfg["fields"]))]

    def _data_rows(self, ws, cfg):
        start, end = cfg["header_row"] + 1, ws.max_row
        rows = []
        for r in range(start, end + 1):
            values = [ws.cell(r, cfg["start_col"] + i).value for i in range(len(cfg["fields"]))]
            if any(v not in (None, "") for v in values): rows.append((r, values))
        return rows

    def read_entries(self, module, presales=None, month=None, limit=50):
        cfg = MODULES[module]; wb = self._load(); ws = wb[cfg["sheet"]]
        items = []
        for row_num, values in reversed(self._data_rows(ws, cfg)):
            item = {field: values[i] for i, field in enumerate(cfg["fields"])}; item["_row"] = row_num
            if presales and str(item.get(cfg["presales"], "")).casefold() != presales.casefold(): continue
            if month and str(item.get("Month", "")).casefold() != month.casefold(): continue
            for k, v in list(item.items()):
                if isinstance(v, datetime): item[k] = v.date().isoformat()
                elif isinstance(v, date): item[k] = v.isoformat()
            items.append(item)
            if len(items) >= min(limit, 5000): break
        wb.close(); return items

    def _find_insert_row(self, ws, cfg):
        start = cfg["header_row"] + 1
        if cfg["sheet"] == "Win-Lost":
            for r in range(start, ws.max_row + 2):
                joined = " ".join(str(ws.cell(r, c).value or "") for c in range(1, ws.max_column + 1)).casefold()
                if "total po received" in joined or any(ws.cell(r, c).data_type == "f" for c in range(1, ws.max_column + 1)):
                    ws.insert_rows(r)
                    # openpyxl deliberately does not update dependent formulas on row
                    # insertion. Extend summary ranges that ended on the former last
                    # data row so the newly inserted record is included.
                    for rr in range(r + 1, ws.max_row + 1):
                        for cell in ws[rr]:
                            if cell.data_type == "f" and isinstance(cell.value, str):
                                cell.value = re.sub(rf"(:\$?[A-Z]+\$?){r-1}(?=[,)])", rf"\g<1>{r}", cell.value)
                    return r
        rows = self._data_rows(ws, cfg)
        return (rows[-1][0] + 1) if rows else start

    def _backup(self):
        BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        shutil.copy2(self.path, BACKUP_DIR / f"Presales_Weekly_Tracker_{stamp}.xlsx")
        backups = sorted(BACKUP_DIR.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
        for old in backups[10:]: old.unlink()

    def _atomic_save(self, wb):
        fd, name = tempfile.mkstemp(suffix=".xlsx", dir=self.path.parent); os.close(fd); tmp = Path(name)
        try:
            wb.save(tmp); check = openpyxl.load_workbook(tmp, read_only=True); check.close(); os.replace(tmp, self.path)
        finally:
            if tmp.exists(): tmp.unlink()

    def _write(self, operation):
        self.path.parent.mkdir(parents=True, exist_ok=True)
        with self._thread_lock, portalocker.Lock(str(self.lock_path), timeout=30):
            wb = self._load(); self._backup()
            try: result = operation(wb); self._atomic_save(wb); return result
            finally: wb.close()

    def append_entry(self, module: str, data: dict[str, Any]):
        cfg = MODULES[module]
        unknown = set(data) - set(cfg["fields"])
        if unknown: raise ValueError(f"Unsupported fields: {', '.join(sorted(unknown))}")
        def op(wb):
            ws = wb[cfg["sheet"]]; row = self._find_insert_row(ws, cfg); self._copy_style(ws, row - 1, row, cfg["start_col"], len(cfg["fields"]))
            if cfg["serial"]:
                idx = cfg["fields"].index(cfg["serial"]); nums = []
                for _, vals in self._data_rows(ws, cfg):
                    try: nums.append(int(vals[idx]))
                    except (TypeError, ValueError): pass
                data[cfg["serial"]] = max(nums, default=0) + 1
            for i, field in enumerate(cfg["fields"]):
                cell = ws.cell(row, cfg["start_col"] + i, data.get(field))
                if cell.value in (None, ""):
                    cell.fill = copy(MISSING_VALUE_FILL)
                elif cell.fill.fill_type == "solid" and cell.fill.fgColor.rgb == MISSING_VALUE_COLOR:
                    cell.fill = PatternFill()
            return {"row": row, "message": "Entry added successfully"}
        return self._write(op)

    def update_entry(self, module: str, row: int, data: dict[str, Any]):
        cfg = MODULES[module]
        unknown = set(data) - set(cfg["fields"])
        if unknown:
            raise ValueError(f"Unsupported fields: {', '.join(sorted(unknown))}")

        def op(wb):
            ws = wb[cfg["sheet"]]
            valid_rows = {row_number for row_number, _ in self._data_rows(ws, cfg)}
            if row not in valid_rows:
                raise ValueError("Workbook row is not an editable data row")
            for i, field in enumerate(cfg["fields"]):
                cell = ws.cell(row, cfg["start_col"] + i)
                if field != cfg["serial"]:
                    cell.value = data.get(field)
                if cell.value in (None, ""):
                    cell.fill = copy(MISSING_VALUE_FILL)
                elif cell.fill.fill_type == "solid" and cell.fill.fgColor.rgb == MISSING_VALUE_COLOR:
                    cell.fill = PatternFill()
            return {"row": row, "message": "Entry updated successfully"}

        return self._write(op)

    @staticmethod
    def _status_layout(ws):
        for row in range(1, min(ws.max_row, 25) + 1):
            headers = {
                str(ws.cell(row, col).value).strip(): col
                for col in range(1, ws.max_column + 1)
                if ws.cell(row, col).value is not None
            }
            if "Presales Name" in headers:
                return row, headers
        raise ValueError("Status sheet is missing the 'Presales Name' header")

    def status(self):
        wb = self._load(); ws = wb["Status"]; header_row, headers = self._status_layout(ws)
        rows = []
        for r in range(header_row + 1, ws.max_row + 1):
            name = ws.cell(r, headers.get("Presales Name", 1)).value
            if name: rows.append({h: ws.cell(r, c).value for h, c in headers.items()})
        wb.close(); return rows

    def update_status(self, presales, module, status, up_to=None):
        label = STATUS_FIELDS[module]
        def op(wb):
            ws = wb["Status"]; header_row, headers = self._status_layout(ws)
            row = next((r for r in range(header_row + 1, ws.max_row + 1) if str(ws.cell(r, headers["Presales Name"]).value).strip().casefold() == presales.strip().casefold()), None)
            if not row: raise ValueError("Presales user not found")
            ws.cell(row, headers[label], status)
            if "Up to" in headers: ws.cell(row, headers["Up to"], up_to or datetime.now().strftime("%d %b %Y"))
            return {"message": f"{label} marked {status}"}
        return self._write(op)

    def users(self): return [str(r["Presales Name"]) for r in self.status()]

    def raw_sheet(self, name, header_row=1):
        wb = self._load(); ws = wb[name]; headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
        rows = [{str(headers[i] or f"Column {i+1}"): ws.cell(r, i+1).value for i in range(len(headers))} for r in range(header_row+1, ws.max_row+1) if any(ws.cell(r, c).value is not None for c in range(1, ws.max_column+1))]
        wb.close(); return rows

    def dashboard(self, presales):
        counts = {key: len(self.read_entries(key, presales, limit=100)) for key in MODULES}
        status = next((r for r in self.status() if str(r.get("Presales Name", "")).casefold() == presales.casefold()), {})
        recent = []
        for key in MODULES:
            for item in self.read_entries(key, presales, limit=3): recent.append({"module": key, **item})
        return {"counts": counts, "status": status, "recent": recent[:10]}
