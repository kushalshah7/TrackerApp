from __future__ import annotations

from copy import copy
from datetime import date
from io import BytesIO
from typing import Any

import openpyxl
from openpyxl.styles import PatternFill

from .config import MODULES
from .excel_service import ExcelService, MISSING_VALUE_COLOR, MISSING_VALUE_FILL


def is_summary_row(ws, row: int) -> bool:
    text = " ".join(str(ws.cell(row, col).value or "") for col in range(1, ws.max_column + 1)).casefold()
    return "total po received" in text or any(ws.cell(row, col).data_type == "f" for col in range(1, ws.max_column + 1))


def editable_rows(ws, cfg: dict[str, Any]) -> list[int]:
    rows: list[int] = []
    for row in range(cfg["header_row"] + 1, ws.max_row + 1):
        if cfg["sheet"] == "Win-Lost" and is_summary_row(ws, row):
            break
        values = [ws.cell(row, cfg["start_col"] + index).value for index in range(len(cfg["fields"]))]
        if any(value not in (None, "") for value in values):
            rows.append(row)
    return rows


def _copy_row_style(ws, source: int, target: int, cfg: dict[str, Any]) -> None:
    ws.row_dimensions[target].height = ws.row_dimensions[source].height
    for col in range(cfg["start_col"], cfg["start_col"] + len(cfg["fields"])):
        before, after = ws.cell(source, col), ws.cell(target, col)
        if before.has_style:
            after.font = copy(before.font)
            after.fill = copy(before.fill)
            after.border = copy(before.border)
            after.alignment = copy(before.alignment)
            after.protection = copy(before.protection)
            after.number_format = before.number_format


def _write_module(ws, cfg: dict[str, Any], records: list[dict[str, Any]]) -> None:
    start = cfg["header_row"] + 1
    existing = editable_rows(ws, cfg)
    summary_row = None
    if cfg["sheet"] == "Win-Lost":
        summary_row = next((row for row in range(start, ws.max_row + 1) if is_summary_row(ws, row)), ws.max_row + 1)

    occupied: set[int] = set()
    positioned: list[tuple[int, dict[str, Any], bool]] = []
    next_row = max(existing, default=start - 1) + 1
    if cfg["sheet"] == "Win-Lost" and not existing:
        next_row = start
    for record in records:
        source_row = record.get("_source_row")
        preserved = isinstance(source_row, int) and source_row >= start and source_row not in occupied
        if preserved and summary_row is not None and source_row >= summary_row:
            preserved = False
        row = source_row if preserved else next_row
        while row in occupied:
            row += 1
        positioned.append((row, record, preserved))
        occupied.add(row)
        if not preserved:
            next_row = row + 1

    required_end = max(occupied, default=start - 1)
    if summary_row is not None:
        while required_end >= summary_row:
            ws.insert_rows(summary_row)
            _copy_row_style(ws, max(start, summary_row - 1), summary_row, cfg)
            for target_row in range(summary_row + 1, ws.max_row + 1):
                for cell in ws[target_row]:
                    if cell.data_type == "f" and isinstance(cell.value, str):
                        cell.value = cell.value.replace(f":G{summary_row - 1}", f":G{summary_row}")
            positioned = [(row + 1 if row >= summary_row else row, record, preserved) for row, record, preserved in positioned]
            summary_row += 1

    for row in existing:
        for index in range(len(cfg["fields"])):
            ws.cell(row, cfg["start_col"] + index).value = None
    for row, _, preserved in positioned:
        if not preserved and row not in existing:
            _copy_row_style(ws, max(start, row - 1), row, cfg)

    for row, record, preserved in positioned:
        for index, field in enumerate(cfg["fields"]):
            cell = ws.cell(row, cfg["start_col"] + index)
            value = record.get(field)
            if isinstance(value, str) and field in {"Date", "PO Date", "Expected Completion Date"}:
                try:
                    value = date.fromisoformat(value)
                except ValueError:
                    pass
            cell.value = value
            if not preserved:
                if value in (None, ""):
                    cell.fill = copy(MISSING_VALUE_FILL)
                elif cell.fill.fill_type == "solid" and cell.fill.fgColor.rgb == MISSING_VALUE_COLOR:
                    cell.fill = PatternFill()


def build_workbook(template: bytes, entries: dict[str, list[dict[str, Any]]], statuses: list[dict[str, Any]]) -> bytes:
    workbook = openpyxl.load_workbook(BytesIO(template), data_only=False)
    for module, cfg in MODULES.items():
        _write_module(workbook[cfg["sheet"]], cfg, entries.get(module, []))
    status_ws = workbook["Status"]
    header_row, headers = ExcelService._status_layout(status_ws)
    existing_end = max((row for row in range(header_row + 1, status_ws.max_row + 1) if status_ws.cell(row, headers["Presales Name"]).value), default=header_row)
    for row in range(header_row + 1, max(existing_end, header_row + len(statuses)) + 1):
        for column in headers.values():
            status_ws.cell(row, column).value = None
    for offset, status in enumerate(statuses):
        row = header_row + 1 + offset
        for heading, column in headers.items():
            status_ws.cell(row, column).value = status.get(heading)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    result = output.getvalue()
    check = openpyxl.load_workbook(BytesIO(result), data_only=False)
    for cfg in MODULES.values():
        found = [check[cfg["sheet"]].cell(cfg["header_row"], cfg["start_col"] + index).value for index in range(len(cfg["fields"]))]
        if [str(value).rstrip() for value in found] != [field.rstrip() for field in cfg["fields"]]:
            check.close()
            raise ValueError(f"Export header validation failed for {cfg['sheet']}")
    check.close()
    return result

