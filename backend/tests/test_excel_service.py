from pathlib import Path
from datetime import datetime
import openpyxl
import pytest
from app.config import MODULES
from app.excel_service import ExcelService

@pytest.fixture()
def service(tmp_path, monkeypatch):
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    for cfg in MODULES.values():
        ws = wb.create_sheet(cfg["sheet"]); start = cfg["start_col"]
        for i, field in enumerate(cfg["fields"]): ws.cell(cfg["header_row"], start+i, field)
    status = wb.create_sheet("Status")
    status.append(["Presales Name", "Weekly Review", "Weekly Meeting", "Training Attended", "Training Conducted", "PoC Tracker", "Customer Workshop", "Up to"])
    status.append(["Aditya"] + ["Pending"] * 6 + [""])
    wb.create_sheet("Client Managers Target").append(["Region", "AM Name", "Q1", "Q2", "Q3", "Q4"])
    wb.create_sheet("Column Guide").append(["Column", "Explanation"])
    path = tmp_path / "tracker.xlsx"; wb.save(path)
    import app.excel_service as es
    monkeypatch.setattr(es, "BACKUP_DIR", tmp_path / "backups")
    return ExcelService(path)

def value_for(field):
    if field in {"Region"}: return "West"
    if field in {"Presales", "PreSales Name"}: return "Aditya"
    if "Date" in field or field == "PO Date": return "2026-08-25"
    if field in {"Value (₹)", "Deal Value"}: return 1000
    if field.startswith("Sr"): return None
    return "Test"

def test_all_modules_append_and_reopen(service):
    for module, cfg in MODULES.items():
        data = {f: value_for(f) for f in cfg["fields"] if not f.startswith("Sr")}
        service.append_entry(module, data)
        assert service.read_entries(module, "Aditya")[0][cfg["presales"]] == "Aditya"
    wb = openpyxl.load_workbook(service.path)
    assert set(cfg["sheet"] for cfg in MODULES.values()).issubset(wb.sheetnames)
    wb.close()
    assert len(list((service.path.parent / "backups").glob("*.xlsx"))) > 0

def test_status_update(service):
    service.update_status("Aditya", "weekly-review", "Completed", "Week 4, August 2026")
    assert service.status()[0]["Weekly Review"] == "Completed"

def test_unfilled_columns_are_red_without_leaking_to_filled_cells(service):
    data = {"Region": "West", "Presales": "Aditya", "Customer": "Example", "Opportunity Details": "Initial scope"}
    first_row = service.append_entry("weekly-review", data)["row"]
    second_row = service.append_entry("weekly-review", {**data, "AM": "Manager"})["row"]

    wb = openpyxl.load_workbook(service.path)
    ws = wb["Weekly Review"]
    am_col = MODULES["weekly-review"]["fields"].index("AM") + 1
    assert ws.cell(first_row, am_col).fill.fgColor.rgb == "FFFF0000"
    assert ws.cell(second_row, am_col).value == "Manager"
    assert ws.cell(second_row, am_col).fill.fgColor.rgb != "FFFF0000"
    wb.close()

def test_completely_blank_entry_is_allowed_and_empty_cells_are_red(service):
    row = service.append_entry("weekly-meeting", {})["row"]

    wb = openpyxl.load_workbook(service.path)
    cfg = MODULES["weekly-meeting"]
    ws = wb[cfg["sheet"]]
    first_optional_column = cfg["start_col"] + (1 if cfg["serial"] else 0)
    for column in range(first_optional_column, cfg["start_col"] + len(cfg["fields"])):
        cell = ws.cell(row, column)
        assert cell.value is None
        assert cell.fill.fgColor.rgb == "FFFF0000"
    wb.close()

def test_entry_can_be_updated_and_reopened(service):
    data = {"Region": "West", "Presales": "Aditya", "Customer": "Example", "Opportunity Details": "Initial scope"}
    row = service.append_entry("weekly-review", data)["row"]
    service.update_entry("weekly-review", row, {**data, "Customer": "Updated customer", "OEM": "Example OEM"})

    item = service.read_entries("weekly-review", limit=100)[0]
    assert item["Customer"] == "Updated customer"
    assert item["OEM"] == "Example OEM"

def test_datetime_values_are_returned_without_time(service):
    data = {"Region": "West", "Presales": "Aditya", "Account Name": "Example", "Date": datetime(2026, 8, 25, 14, 30)}
    service.append_entry("weekly-meeting", data)
    assert service.read_entries("weekly-meeting", limit=100)[0]["Date"] == "2026-08-25"
