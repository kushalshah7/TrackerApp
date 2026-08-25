from copy import copy
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

import openpyxl
import pytest

from app.config import MODULES
from app.excel_service import ExcelService
from app.workbook_export import build_workbook, editable_rows, is_summary_row


TEMPLATE = Path(__file__).parents[1] / "data" / "Workbook.xlsx"


def workbook_records(path: Path):
    workbook = openpyxl.load_workbook(path, data_only=False)
    entries = {}
    for module, cfg in MODULES.items():
        ws = workbook[cfg["sheet"]]
        entries[module] = [
            {
                **{field: ws.cell(row, cfg["start_col"] + index).value for index, field in enumerate(cfg["fields"])},
                "_source_row": row,
            }
            for row in editable_rows(ws, cfg)
        ]
    statuses = ExcelService(path).status()
    workbook.close()
    return entries, statuses


@pytest.mark.skipif(not TEMPLATE.exists(), reason="private director workbook is not present")
def test_export_round_trip_preserves_director_workbook_structure():
    entries, statuses = workbook_records(TEMPLATE)
    exported = build_workbook(TEMPLATE.read_bytes(), entries, statuses)
    with ZipFile(BytesIO(exported)) as archive:
        assert archive.testzip() is None
    before = openpyxl.load_workbook(TEMPLATE, data_only=False)
    after = openpyxl.load_workbook(BytesIO(exported), data_only=False)
    assert after.sheetnames == before.sheetnames
    for sheet_name in before.sheetnames:
        source, result = before[sheet_name], after[sheet_name]
        assert {str(item) for item in result.merged_cells.ranges} == {str(item) for item in source.merged_cells.ranges}
        assert result.sheet_view.showGridLines == source.sheet_view.showGridLines
        assert result.freeze_panes == source.freeze_panes
        for column, dimension in source.column_dimensions.items():
            assert result.column_dimensions[column].width == dimension.width
        for source_row in source.iter_rows():
            for source_cell in source_row:
                result_cell = result[source_cell.coordinate]
                assert result_cell.value == source_cell.value
                assert copy(result_cell.font) == copy(source_cell.font)
                assert copy(result_cell.fill) == copy(source_cell.fill)
                assert copy(result_cell.border) == copy(source_cell.border)
                assert copy(result_cell.alignment) == copy(source_cell.alignment)
                assert result_cell.number_format == source_cell.number_format
                assert copy(result_cell.protection) == copy(source_cell.protection)
        source_formulas = {cell.coordinate: cell.value for row in source.iter_rows() for cell in row if cell.data_type == "f"}
        result_formulas = {cell.coordinate: cell.value for row in result.iter_rows() for cell in row if cell.data_type == "f"}
        assert result_formulas == source_formulas
    before.close()
    after.close()


@pytest.mark.skipif(not TEMPLATE.exists(), reason="private director workbook is not present")
def test_export_keeps_summary_out_of_records_and_preserves_typed_values():
    entries, statuses = workbook_records(TEMPLATE)
    assert entries["win-lost"] == []
    entries["win-lost"].append({
        "PO Date": "2026-08-25", "Region": "West", "Presales": "Aditya",
        "Account Name": "Director QA", "Win/Lost": "Win", "Deal Value": 1250000,
        "Remark": "Verified export",
    })
    exported = build_workbook(TEMPLATE.read_bytes(), entries, statuses)
    workbook = openpyxl.load_workbook(BytesIO(exported), data_only=False)
    ws = workbook["Win-Lost"]
    assert ws["B4"].value.year == 2026
    assert ws["G4"].value == 1250000
    summary = next(row for row in range(4, ws.max_row + 1) if is_summary_row(ws, row))
    assert ws.cell(summary, 6).value == "Total PO Received"
    assert ws.cell(summary, 7).data_type == "f"
    workbook.close()

